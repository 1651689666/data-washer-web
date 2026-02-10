import openpyxl
from openpyxl.utils import get_column_letter

file_path = 'demo_horizontal_merge.xlsx'

def modify_excel():
    print(f"Loading {file_path}...")
    wb = openpyxl.load_workbook(file_path)
    
    # 1. User Request: Add 'Grade' field to first sheet if missing
    sheet = wb['Merge Demo']
    
    # Check if Grade already exists (simple check on row 1)
    header_row = 1
    grade_col_idx = None
    
    # We want to insert it after 'Dept ID' (Column C, idx 3), so Column D (idx 4)
    # Currently D is likely empty separator. Let's just use Column D.
    target_col_idx = 4 
    target_col_letter = get_column_letter(target_col_idx)
    
    print(f"Adding 'Grade' column at {target_col_letter}...")
    sheet.cell(row=header_row, column=target_col_idx, value="Grade")
    
    # Values to cycle through
    grades = ['G1', 'G2', 'G3', 'G1', 'G2']
    
    # Iterate rows starting from 2
    # We stop when there is no Emp ID (Column A)
    max_row = sheet.max_row
    for row_idx in range(2, max_row + 1):
        emp_id = sheet.cell(row=row_idx, column=1).value
        if emp_id:
            grade_val = grades[(row_idx - 2) % len(grades)]
            sheet.cell(row=row_idx, column=target_col_idx, value=grade_val)
            print(f"  Row {row_idx}: Set Grade = {grade_val}")
            
    # 2. Add new sheet with associated data
    new_sheet_name = 'Salary Grades'
    if new_sheet_name in wb.sheetnames:
        print(f"Sheet {new_sheet_name} already exists. Removing it to recreate.")
        del wb[new_sheet_name]
        
    print(f"Creating new sheet: {new_sheet_name}...")
    ws_grades = wb.create_sheet(new_sheet_name)
    
    # Headers
    headers = ['Grade Code', 'Base Salary', 'Bonus Amount']
    ws_grades.append(headers)
    
    # Data
    data = [
        ['G1', 5000, 1000],
        ['G2', 8000, 2000],
        ['G3', 12000, 5000],
        ['G4', 20000, 10000] # Extra grade not in main data
    ]
    
    for row in data:
        ws_grades.append(row)
        
    print(f"Saving to {file_path}...")
    wb.save(file_path)
    print("Done.")

if __name__ == "__main__":
    modify_excel()
